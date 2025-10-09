# utils/export.py
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

def render_to_pdf(template_path, context, filename="report.pdf"):
    html = render_to_string(template_path, context)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)

    if pdf.err:
        return HttpResponse("Error rendering PDF", status=500)

    response = HttpResponse(result.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def render_to_excel(data_dict, filename="report.xlsx"):
    wb = openpyxl.Workbook()

    header_fill = PatternFill(start_color="FFDDEBF7", end_color="FFDDEBF7", fill_type="solid")
    bold_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for sheet_name, rows in data_dict.items():
        ws = wb.create_sheet(title=sheet_name[:31])

        # Landscape layout
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)
        ws.sheet_properties = WorksheetProperties(pageSetUpPr=PageSetupProperties(fitToPage=True))
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)

            # Style headers
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = bold_font
                cell.alignment = center
                cell.fill = header_fill
                cell.border = thin_border

            # Body
            for row in rows:
                values = list(row.values())
                ws.append(values)
                row_idx = ws.max_row
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                        cell.alignment = right
                    cell.border = thin_border

            # Freeze header
            ws.freeze_panes = "A2"

            # Auto-size columns
            for col_idx in range(1, len(headers) + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[col_letter]:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        if len(val) > max_length:
                            max_length = len(val)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max(12, max_length + 2), 50)

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    response = HttpResponse(
        file_stream.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
